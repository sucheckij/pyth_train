import json
import os
import pickle
import shelve
# openpyxl - biblioteka do excela w Pythonie
from xml.etree.ElementTree import *
import openpyxl
import shutil


def add_car(root, rej, brand, model, color, year, price, currency, motor_type):
    car = SubElement(root,'car',rej=rej)
    SubElement(car,'brand').text= brand
    SubElement(car,'model').text= model
    SubElement(car,'color').text= color
    SubElement(car,'year').text= f"{year}"
    SubElement(car,'price', currency=currency).text= str(price)
    SubElement(car,'motor_type').text= motor_type





def demo_pliki():
    print("Praca z plikami")
    print("1 - odczyt.txt")
    print("2 - zapis.xml")
    print("3 - odczyt.xml")
    print("4 - update.xml")
    print("5 - delete from .xml")
    print("6 - pickle")
    print("7 - shelve")
    print("8 - json")
    print("9 - odczyt.txt")
    print("10- archiwizacja zip i tar")

    odp = input("Co robimy?")
    if odp == "1":
        with open(f"files/raven.txt", "r", encoding="utf-8") as f:
            print(f.read())
            f.seek(0)
            print(f.readline(20))
            print(f.readline())
            f.seek(0)
            lista = f.readlines()
            print(len(lista))
            print(lista)
    elif odp == "2":
        root = Element('cars')
        rej, brand, model, color, year, prize, currency, motor_type = 'DW12345', 'BMW', 'Serie 4', 'czarny',2007,115, 'PLN', 'benzyna'
        add_car(root, rej, brand,model,color, year,prize,currency,motor_type)
        rej, brand, model, color, year, prize, currency, motor_type = 'ONA2345', 'Trabant', 'Noname', 'zielen', 1960, 100, 'EUR', 'benzyna'
        add_car(root, rej, brand, model, color, year, prize, currency, motor_type)
        ElementTree(root).write(r'files/data.xml',encoding="utf-8",xml_declaration=True)

    elif odp == "3":
        p = parse(r"files/data.xml")
        r = p.getroot()
        for i in r.findall('car'):
            marka = i.find('brand').text
            model = i.find("model").text
            cena = i.find("price")
            wartosc = cena.text
            waluta = cena.get('currency')

        print(marka,model,wartosc,waluta)
    elif odp == "4":
        p = parse(r"files\cars.xml")
        r = p.getroot()
        szuk_rej = input("Nr rej auta -> ")
        for c in r.findall('car'):
            if c.get('rej') == szuk_rej:
                cena = input('Podaj aktualną cenę -> ')
                cd = c.find('price')
                cd.text = cena
    elif odp == "5":
        p = parse(r"files\cars.xml")
        r = p.getroot()
        szuk_rej = input("Nr rej auta -> ")
        for c in r.findall('car'):
            if c.get('rej') == szuk_rej:
                r.remove(c)

        ElementTree(r).write(r"files\cars.xml", encoding= "uft-8",xml_declaration=True)
    elif odp == "6":
        print("marynowanie danych - pickle")
        przepis = ["kiszenie","marynowanie"]
        shape = ["cale", "cwiarki", "plasterki"]
        made_by = ["pudliszki","rolnik"]
        file = open(r"files/marynaty.dat","wb")
        pickle.dump(przepis,file)
        pickle.dump(shape,file)
        pickle.dump(made_by,file)
        file.close()

        print("Odczyt danych")  # serializacja to proces tworzenia struktury danych w odpowiedniej postaci
        plik_p = open(r"files/marynaty.dat", "rb")
        r = pickle.load(plik_p)
        s = pickle.load(plik_p)
        p = pickle.load(plik_p)
        plik_p.close()
        print(r,s,p)


    elif odp == "7":
        print("marynowanie danych - shelve")
        file = shelve.open(r"files/mary", "rb")
        file["przepis"] = ["kiszenie", "marynowanie"]
        file["shape"] = ["cale", "cwiarki", "plasterki"]
        file["made_by"] = ["pudliszki", "rolnik"]
        file.sync()
        file.close()

        print("Odczyt danych")
        p_s = shelve.open(r'file\maryn')
        print(p_s["shape"])
        p_s.close()


    elif odp == "8":  # JSON
        config = {
            "server" : "(local)",
            "database" : "DbMichal",
            "uid": "michal",
            "password": "Password"
        }

        with open(r"files/json_example", 'w') as f:
            json.dump(config,f)

        with open(r"files/json_example", 'r') as file:
            dane = json.load(file)
            print(dane["server"])

        pass
    elif odp == "9":
        wb = openpyxl.Workbook()
        wsn = wb.create_sheet('Auta')
        wsn['B2']= '=sum(1,2,3,4)'
        wb.save(r"files/auta.xlsx")

        wbr = openpyxl.load_workbook(filename=r"files/auta.xlsx")
        r = wbr['Auta']
        print(r['B2'].value)

    elif odp == '10':
        shutil.make_archive(rf"{os.getcwd()}\archiwum\pliki",format='zip',root_dir=r'files')
        shutil.make_archive(rf"{os.getcwd()}\archiwum\pliki",format='tar',root_dir=r'files')
